import openpyxl
import os

file_path = "Infill New.xlsx"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    exit()

print(f"Inspecting {file_path}...")
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

print(f"Max Rows: {sheet.max_row}")
print(f"Max Cols: {sheet.max_column}")

for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5)):
    row_num = i + 1
    print(f"\n--- Row {row_num} ---")
    cells = [c for c in row]
    
    # Print first few cells detailed info
    for j, cell in enumerate(cells[:5]):
        val = cell.value
        is_bold = cell.font.b if cell.font else False
        print(f"  Col {j} ('{val}'): Bold={is_bold}")

    # Check header candidates
    values = [str(c.value).strip() if c.value else "" for c in row]
    if 'Task' in values:
        print(f"  -> HEADER DETECTED? 'Task' found at index {values.index('Task')}")
